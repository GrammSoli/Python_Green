import os
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from selenium import webdriver
from bs4 import BeautifulSoup
import datetime
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


# Шаг 1: Выбор даты
def get_previous_month_range(date):
    # Вычисляем первый день текущего месяца
    first_day_of_month = datetime.date(date.year, date.month, 1)

    # Вычисляем первый день предыдущего месяца
    last_month = date.month - 1 if date.month > 1 else 12
    last_month_year = date.year if date.month > 1 else date.year - 1
    first_day_of_last_month = datetime.date(last_month_year, last_month, 1)

    # Вычисляем последний день предыдущего месяца
    last_day_of_last_month = first_day_of_month - datetime.timedelta(days=1)

    # Получаем только день первого и последнего дня предыдущего месяца
    first_day = first_day_of_last_month.day
    last_day = last_day_of_last_month.day

    # Вычисляем название предыдущего месяца
    previous_month = str(last_month)

    return first_day, last_day, previous_month


today = datetime.date.today() # Получаем сегодняшнюю дату

first_day, last_day, previous_month = get_previous_month_range(today) # Получаем первый и последний день предыдущего месяца для сегодняшней даты


currency = ["USD_RUB", "JPY_RUB"]

# Удаление файла, если он существует
if os.path.exists('data.xlsx'):
    os.remove('data.xlsx')

df = pd.DataFrame()

for n in currency:

    # Установка пути к веб-драйверу (например, Chrome)
    driver_path = 'C:/Users/Mikhail/Documents/chromedriver.exe'

    # Создание экземпляра веб-драйвера
    driver = webdriver.Chrome(driver_path)

    # Шаг 1: Открыть https://www.moex.com
    driver.get(f"https://www.moex.com/ru/derivatives/currency-rate.aspx?currency={n}")

    #  ожидание загрузки страницы
    wait = WebDriverWait(driver, 10)
    wait.until(EC.title_contains("Индикативные курсы валют — Московская Биржа | Рынки"))

    # Соглашение
    wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="content_disclaimer"]/div/div/div/div[1]/div/a[1]')))
    soglasie_xpath = '//*[@id="content_disclaimer"]/div/div/div/div[1]/div/a[1]'
    soglasie_link = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, soglasie_xpath)))
    soglasie_link.click()

    # Меняем числа
    def change_dropdown(driver, dropdown_xpath, dropdown_id, value):
        dropdown = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, dropdown_xpath)))
        script = f"document.getElementById('{dropdown_id}').value = '{value}';"
        driver.execute_script(script)

    change_dropdown(driver, '//*[@id="d1day"]', 'd1day', first_day)
    change_dropdown(driver, '//*[@id="d2day"]', 'd2day', last_day)
    change_dropdown(driver, '//*[@id="d1month"]', 'd1month', previous_month)
    change_dropdown(driver, '//*[@id="d2month"]', 'd2month', previous_month)

    # Нажатие кнопки "Показать"
    pokazat_xpath = '//*[@value="Показать"]'
    pokazat_link = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, pokazat_xpath)))
    pokazat_link.click()

    # Получение HTML-кода страницы
    html = driver.page_source

    # Создание объекта BeautifulSoup для парсинга HTML
    soup = BeautifulSoup(html, 'html.parser')

    # Нахождение всех таблиц на странице
    tables = soup.find_all('table')

    # Проверка, что на странице есть хотя бы две таблицы
    if len(tables) >= 2:
        # Выбор таблицы
        table = tables[4]

        # Получение заголовков таблицы
        headers = table.find_all('th')
        header_row = [header.get_text(strip=True) for header in headers]

        # Удаление ненужных столбцов из заголовков
        del header_row[1:5]

        # Создание пустого DataFrame с заданными столбцами
        df = pd.DataFrame(columns=header_row)

        # Получение строк таблицы
        rows = table.find_all('tr')
        skip_rows = False  # Флаг для пропуска строк с данными

        for row in rows:
            # Получение ячеек в текущей строке
            cells = row.find_all('td')

            # Создание копии header_row для удаления столбцов
            temp_header_row = header_row.copy()

            # Удаление ненужных ячеек
            if len(cells) > 2:
                del cells[2]
            if len(cells) > 1:
                del cells[1]

            # Получение данных из ячеек текущей строки
            row_data = [cell.get_text(strip=True) for cell in cells]

            # Удаление ненужных столбцов из данных текущей строки
            # del row_data[1:3]

            # Проверка длины row_data и temp_header_row
            if len(row_data) == len(temp_header_row):
                # Создание нового DataFrame с временными столбцами
                temp_df = pd.DataFrame([row_data], columns=temp_header_row)
                # Объединение временного DataFrame с основным DataFrame df
                df = pd.concat([temp_df, df], ignore_index=True)  # Замена df на temp_df и изменение порядка объединения


        # Проверка наличия файла 'data.xlsx'
    if os.path.isfile('data.xlsx'):
        existing_data = pd.read_excel('data.xlsx', sheet_name='Sheet1')
    else:
        existing_data = pd.DataFrame()  # Создание пустого DataFrame, если файл не найден

        # Определение начального столбца для добавления новых данных
    start_column = existing_data.shape[1] if not existing_data.empty else 0

        # Объединение существующих данных и новых данных вместе
    merged_df = pd.concat([existing_data, df], axis=1)


        # Сохранение объединенного DataFrame в Excel
    excel_writer = pd.ExcelWriter('data.xlsx', engine='openpyxl')
    merged_df.to_excel(excel_writer, sheet_name='Sheet1', index=False)
    excel_writer.book.save('data.xlsx')
    excel_writer.book.close()



    # Закрытие браузера
    driver.quit()


# Работа с Excel

workbook = openpyxl.load_workbook('data.xlsx')

# Выбираем активный лист
sheet = workbook.active

# Получаем максимальное количество строк в файле
max_row = sheet.max_row

# Вносим заголовок столбца G
sheet['G1'] = 'Результат'

# Проходим по каждой строке и выполняем деление
for i in range(2, max_row + 1):
    value_b = float(sheet[f'B{i}'].value.replace(',', '.'))
    value_e = float(sheet[f'E{i}'].value.replace(',', '.'))

    if value_e != 0:
        result = value_b / value_e
        sheet[f'G{i}'].value = result

# Выравнивание автоширины
for column_cells in sheet.columns:
    max_length = 0
    column = column_cells[0].column_letter
    for cell in column_cells:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[column].width = adjusted_width

# Установка финансового формата чисел
number_format = '0.00 ₽'  # Формат чисел в рублях
for row in sheet.iter_rows(min_row=2, min_col=7, max_col=7):
    for cell in row:
        cell.number_format = number_format

# Сохраняем изменения
workbook.save('data.xlsx')
workbook.close()


# Склонение строк
if max_row == 1:
    rowCountString = str(max_row) + " строка"
elif (max_row % 10 == 1 or max_row % 100 == 11) and max_row != 111:
    rowCountString = str(max_row) + " строка"
elif (max_row % 10 >= 2 and max_row % 10 <= 4) and (max_row % 100 < 10 or max_row % 100 >= 20):
    rowCountString = str(max_row) + " строки"
else:
    rowCountString = str(max_row) + " строк"

# Отправка отчета по Email

# Отправитель
sender_email = "u-698@yandex.ru"  # Замените на свой адрес электронной почты
sender_password = "dvxqgsxsijjzzzkp"  # Замените на свой пароль от электронной почты

# Получатель
receiver_email = "wdblackmonster@gmail.com"  # Замените на адрес электронной почты получателя

# Создание объекта сообщения
message = MIMEMultipart()
message["From"] = sender_email
message["To"] = receiver_email
message["Subject"] = "Отчет"

# Текстовое содержимое письма
body = "Добрый день,\n\nВ файле 'data.xlsx' " + rowCountString + "."
message.attach(MIMEText(body, "plain"))

# Прикрепление файла
filename = "data.xlsx"  # Укажите имя вашего файла
attachment = open(filename, "rb")

part = MIMEBase("application", "octet-stream")
part.set_payload((attachment).read())
encoders.encode_base64(part)
part.add_header("Content-Disposition", "attachment", filename=filename)

message.attach(part)

# Установка соединения с SMTP-сервером и отправка письма
try:
    server = smtplib.SMTP_SSL("smtp.yandex.ru", 465)  # Замените на адрес и порт вашего SMTP-сервера
    # server.starttls()
    server.login(sender_email, sender_password)
    server.sendmail(sender_email, receiver_email, message.as_string())
    server.quit()
    print("Письмо успешно отправлено.")
except Exception as e:
    print("Ошибка при отправке письма:", str(e))